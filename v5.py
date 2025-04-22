from P4 import P4, P4Exception
import mysql.connector
import pandas as pd
import json
from difflib import unified_diff
from datetime import datetime
import openpyxl
import getpass
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Connect to the MySQL database
conn = mysql.connector.connect(
    host="do-dmsrep",
    user="dmsrep",
    database="DMS",
    charset='utf8mb4',
    collation='utf8mb4_unicode_ci'
)

cursor = conn.cursor()

# Define baseline values
baseline1=input("Enter the first baseline: ")
baseline2=input("Enter the second baseline: ")

#define environment ID
envId=input("Enter the environment ID: ")

#getting username and password for perforce
p4_user = "z0052jpm"
# p4_password = getpass.getpass("Enter Perforce password: ")

p4_password ="SatyajitGaikwad!90"

# Perforce connection parameters
p4Params = {
    'Client': "DI2INPUN05643WH",
    'Port': "ssl:146.122.98.36:1667",
    'Pass': p4_password,
    'perfUser': p4_user
}

# Perforce connection setup
p4 = P4(client=p4Params['Client'], port=p4Params['Port'], password=p4Params['Pass'])
p4.user = p4Params['perfUser']


#create excel file and setting width of cells 
def create_excel_file(filtered_details):
    """Create initial Excel file with basic information"""
    df = pd.DataFrame(filtered_details)
    # Add empty columns for detailed differences
    df['added_lines'] = ''
    df['removed_lines'] = ''
    # df['revision_pairs'] = ''
    
    excel_filename = f"file_comparison_{baseline1}_to_{baseline2}.xlsx"
    df.to_excel(excel_filename, index=False)
    
    # Adjust column widths
    workbook = openpyxl.load_workbook(excel_filename)
    worksheet = workbook.active
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 50  # filename
    worksheet.column_dimensions['B'].width = 10  # occurrences
    worksheet.column_dimensions['C'].width = 15  # sourcefile_id
    worksheet.column_dimensions['D'].width = 15  # cp_names
    worksheet.column_dimensions['E'].width = 15  # cp_ids
    worksheet.column_dimensions['F'].width = 100  # added_lines
    worksheet.column_dimensions['G'].width = 100  # removed_lines
    
   
    # Set wrap text for all columns
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    
    workbook.save(excel_filename)
    print(f"Initial Excel file '{excel_filename}' created successfully!")
    return excel_filename


def update_excel_and_json(excel_filename, sourcefile_id, cp_id, modifications, json_data):
    """Update both Excel file and JSON data with differences"""
    # Read existing Excel file
    workbook = openpyxl.load_workbook(excel_filename)
    worksheet = workbook.active
    
    # Find the row with matching sourcefile_id and cp_id
    row_num = None
    for row in range(2, worksheet.max_row + 1):
        if (str(worksheet[f'C{row}'].value) == str(sourcefile_id) and 
            str(worksheet[f'E{row}'].value) == str(cp_id)):
            row_num = row
            break
    
    if row_num:
        # Prepare added lines, removed lines, revision pairs, and detailed differences
        added_lines = []
        removed_lines = []
        revision_pairs = []
        # detailed_differences = []
        
        for mod in modifications:
            rev_pair = f"Rev {mod['from_revision']} → {mod['to_revision']}"
            revision_pairs.append(rev_pair)
            
            diff_lines = mod['differences'].splitlines()
            for line in diff_lines:
                if line.startswith('+'):
                    added_lines.append(line[1:].strip())
                elif line.startswith('-'):
                    removed_lines.append(line[1:].strip())
        
        # Update Excel cells
        worksheet[f'F{row_num}'] = '\n'.join(added_lines)
        worksheet[f'G{row_num}'] = '\n'.join(removed_lines)
        worksheet[f'H{row_num}'] = '\n'.join(revision_pairs)
        # worksheet[f'I{row_num}'] = '\n\n'.join(detailed_differences)
        
    
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col, color in [('F', 'A8DCAB'), ('G', 'f09191')]:
            cell = worksheet[f'{col}{row_num}']
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(color='000000')
            cell.border = thin_border
        
        
        
        # Ensure cells have wrap text enabled
        for col in ['F', 'G']:
            worksheet[f'{col}{row_num}'].alignment = Alignment(wrap_text=True)
        
        # Save the workbook
        workbook.save(excel_filename)
    
    # Update JSON data
    if sourcefile_id in json_data:
        for cp_entry in json_data[sourcefile_id]["cp_entries"]:
            if cp_entry["cp_id"] == cp_id:
                cp_entry["modifications"] = modifications
                break


# Compare two versions of content and return differences
def compare_versions(content1, content2, rev1, rev2):
    """Compare two versions of content and return differences."""
    lines1 = content1.splitlines(keepends=True)
    lines2 = content2.splitlines(keepends=True)
    diff = list(unified_diff(lines1, lines2, fromfile=f'Revision {rev1}', tofile=f'Revision {rev2}'))
    return ''.join(diff) if diff else "No differences found."


# Fetch revisions and compare them using Perforce
def get_revision_and_compare(sourcefile_id, cp_id, file_path, json_data, excel_filename):
    """Fetch revisions and compare them using Perforce"""
    revisions = []
    modifications = []
    
    query = "SELECT revision FROM cpfile WHERE idchangepackage = %s AND idsourcefile = %s AND action='modify' and not type='source' ;"
    cursor.execute(query, (cp_id, sourcefile_id))
    revisions.extend([rev[0] for rev in cursor.fetchall()])


    revisions = sorted(revisions)

    for i in range(len(revisions) - 1):
        rev1 = revisions[i]
        rev2 = revisions[i + 1]

        try:
            content1 = p4.run_print(f"{file_path}#{rev1}")[1]
            content2 = p4.run_print(f"{file_path}#{rev2}")[1]

            differences = compare_versions(content1, content2, rev1, rev2)
            print(f"\nComparing Revision {rev1} with Revision {rev2}:")
            # print(differences)

            modifications.append({
                "from_revision": rev1,
                "to_revision": rev2,
                "differences": differences
            })
        except P4Exception as e:
            print(f"Error fetching content for revisions {rev1} and {rev2}: {e}")

    # Update both Excel and JSON with the modifications
    update_excel_and_json(excel_filename, sourcefile_id, cp_id, modifications, json_data)

# Main script execution
# Connect to Perforce server and fetch baselines within the range
# also to get the chnge packages linked to the baselines and their source files
#
try:
    p4.connect()
    p4.run_login()
    print("Connected to Perforce server.")

    # Query to fetch baselines within the range
    query = """
    SELECT IDbaseline 
    FROM baseline 
    WHERE time > (SELECT time FROM baseline WHERE name = %s) 
    AND time <= (SELECT time FROM baseline WHERE name = %s) 
    AND IDEnvironment = (SELECT IDenvironment FROM environment WHERE envId = %s);
    """

    cursor.execute(query, (baseline1, baseline2, envId))
    baseline_results = cursor.fetchall()

    file_counts = {}
    file_cp_mapping = {}
    file_cp_id_mapping = {}
    file_source_mapping = {}
    filtered_details = []

    count = 0
    for (baseline_id,) in baseline_results:
        # Get change packages linked to this baseline
        query_1 = "SELECT IDchangepackage FROM cpinbaseline WHERE IDbaseline = %s;"
        cursor.execute(query_1, (baseline_id,))
        change_packages = cursor.fetchall()

        for (cp_id,) in change_packages:
            # Get change package name
            get_cp_name = "SELECT name FROM changepackage WHERE IDchangepackage = %s;"
            cursor.execute(get_cp_name, (cp_id,))
            cp_name = cursor.fetchone()

            query_2 = "SELECT distinct idsourcefile FROM cpfile WHERE idchangepackage= %s;"
            cursor.execute(query_2, (cp_id,))
            source_files = cursor.fetchall()

            for (sourcefile_id,) in source_files:
                query_3 = "SELECT name FROM sourcefile WHERE idsourcefile = %s;"
                cursor.execute(query_3, (sourcefile_id,))
                sourcefile_name = cursor.fetchone()
                count += 1
                if sourcefile_name:
                    filename = sourcefile_name[0]
                    file_counts[filename] = file_counts.get(filename, 0) + 1

                    if filename not in file_cp_mapping:
                        file_cp_mapping[filename] = []
                    file_cp_mapping[filename].append(cp_name[0] if cp_name else "Unknown")

                    if filename not in file_cp_id_mapping:
                        file_cp_id_mapping[filename] = []
                    file_cp_id_mapping[filename].append(str(cp_id))

                    if filename not in file_source_mapping:
                        file_source_mapping[filename] = sourcefile_id

    # Add details to filtered_details with separate rows for each CP
    for filename, occurrences in file_counts.items():
        if occurrences > 1:
            sourcefile_id = file_source_mapping[filename]
            cp_names = file_cp_mapping[filename]
            cp_ids = file_cp_id_mapping[filename]
            
            # Create separate row for each CP
            for i in range(len(cp_names)):
                filtered_details.append({
                    "filename": filename,
                    "occurrences": occurrences,
                    "sourcefile_id": sourcefile_id,
                    "cp_names": cp_names[i],  # Single CP name
                    "cp_ids": cp_ids[i]       # Single CP ID
                })

    # Create initial Excel file
    excel_filename = create_excel_file(filtered_details)

    # Create initial JSON data
    json_data = {}
    for detail in filtered_details:
        sourcefile_id = detail['sourcefile_id']
        if sourcefile_id not in json_data:
            json_data[sourcefile_id] = {
                "filename": detail['filename'],
                "occurrences": detail['occurrences'],
                "cp_entries": []
            }
        
        json_data[sourcefile_id]["cp_entries"].append({
            "cp_name": detail['cp_names'],
            "cp_id": detail['cp_ids'],
            "modifications": []
        })

    # Process files and update both Excel and JSON
    print("\nProcessing files and updating documents...")
    # processed_count = 0
    for detail in filtered_details:
        sourcefile_id = detail['sourcefile_id']
        cp_id = detail['cp_ids']
        query_to_root_path = "SELECT rootpath FROM cpfile WHERE idchangepackage = %s AND idsourcefile = %s AND rootpath  like '//plm%';"
        cursor.execute(query_to_root_path, (cp_id, sourcefile_id))
        result = cursor.fetchall()
        rootpath = result[0][0] if result and result[0][0] else ""
        if rootpath:
            rootpath += "/"
        print("rootpath ",result)
        

        file_path = f"{rootpath}{detail['filename']}"
        
        print(f"\nProcessing Source File ID: {sourcefile_id}, CP ID: {cp_id}")
        get_revision_and_compare(sourcefile_id, cp_id, file_path, json_data, excel_filename)
        print("-" * 50)
        # processed_count += 1

    # Write final JSON data
    json_filename = f"{baseline1}_to_{baseline2}.json"
    with open(json_filename, "w", encoding="utf-8") as json_file:
        json.dump(json_data, json_file, indent=4)

    print(f"\nProcessing complete!")
    print(f"Excel file: {excel_filename}")
    print(f"JSON file: {json_filename}")

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    # Close connections
    p4.disconnect()
    print("Disconnected from Perforce server.")
    cursor.close()
    conn.close()